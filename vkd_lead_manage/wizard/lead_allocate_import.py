import base64
import io
from datetime import date

from odoo import models, fields, _
from odoo.exceptions import UserError
from odoo.tools.misc import xlsxwriter


class LeadAllocateImport(models.TransientModel):
    _name = 'lead.allocate.import'
    _description = 'Import Lead Allocations from Excel'

    import_file = fields.Binary(string='Excel File (.xlsx)', attachment=False)
    import_file_name = fields.Char(string='File Name')

    def action_download_template(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        self._build_instructions_sheet(workbook)
        self._build_data_sheet(workbook)
        workbook.close()
        output.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'Lead Allocation Import Template.xlsx',
            'datas': base64.b64encode(output.getvalue()).decode('utf-8'),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d' % attachment.id,
            'target': 'download',
        }

    def _build_instructions_sheet(self, workbook):
        ws = workbook.add_worksheet('Instructions')

        title_fmt = workbook.add_format({
            'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#1F4E79', 'font_color': '#FFFFFF', 'border': 1,
        })
        hdr_fmt = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#2E75B6', 'font_color': '#FFFFFF', 'border': 1,
        })
        cell_fmt = workbook.add_format({
            'align': 'left', 'valign': 'vcenter', 'border': 1, 'text_wrap': True,
        })
        req_fmt = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#FCE4D6', 'font_color': '#C55A11', 'border': 1,
        })
        cond_fmt = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#FFEB9C', 'font_color': '#9C6500', 'border': 1,
        })
        opt_fmt = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#E2EFDA', 'font_color': '#375623', 'border': 1,
        })
        note_fmt = workbook.add_format({
            'bold': True, 'italic': True, 'font_color': '#C00000',
            'text_wrap': True, 'valign': 'vcenter',
        })

        ws.set_column(0, 0, 10)
        ws.set_column(1, 1, 26)
        ws.set_column(2, 2, 16)
        ws.set_column(3, 3, 62)

        ws.merge_range(0, 0, 0, 3, 'Lead Allocation Import — Instructions', title_fmt)
        ws.set_row(0, 32)

        for col, header in enumerate(['Column', 'Field', 'Required', 'Description']):
            ws.write(1, col, header, hdr_fmt)
        ws.set_row(1, 20)

        rows = [
            ('A', 'Team',                 'Required',    'Exact name of the CRM Team (e.g. "Sales Team A")'),
            ('B', 'Distribution Type',    'Required',    'Enter "single" for Single Agent, or "round_robin" for Round Robin'),
            ('C', 'Shift Name',           'Required',    'Exact name of the Shift (e.g. "Morning Shift")'),
            ('D', 'Shift Date',           'Required',    'Date in YYYY-MM-DD format (e.g. 2025-06-01)'),
            ('E', 'Single Agent',         'Conditional', 'Agent display name or login — required when Distribution Type is "single"'),
            ('F', 'RR Agent 1',           'Conditional', 'First agent — required when Distribution Type is "round_robin"'),
            ('G', 'RR Agent 2',           'Optional',    'Second agent for Round Robin (leave blank if not needed)'),
            ('H', 'RR Agent 3',           'Optional',    'Third agent for Round Robin (leave blank if not needed)'),
            ('I', 'RR Agent 4',           'Optional',    'Fourth agent for Round Robin (leave blank if not needed)'),
            ('J', 'RR Agent 5',           'Optional',    'Fifth agent for Round Robin — maximum 5 agents total'),
            ('K', 'Allocation Type',      'Optional',    'Enter "apartment" (default) or "lands". Land leads (adset name containing "land"/"lands") are routed to the Lands allocation'),
        ]
        req_map = {'Required': req_fmt, 'Conditional': cond_fmt, 'Optional': opt_fmt}

        for r, (col, field, req, desc) in enumerate(rows, start=2):
            ws.write(r, 0, col, cell_fmt)
            ws.write(r, 1, field, cell_fmt)
            ws.write(r, 2, req, req_map[req])
            ws.write(r, 3, desc, cell_fmt)
            ws.set_row(r, 18)

        ws.merge_range(13, 0, 14, 3,
            'IMPORTANT NOTES:\n'
            '1. Do NOT modify or delete the column headers in the "Import Data" sheet.\n'
            '2. Delete the grey sample rows before uploading.\n'
            '3. Agent names must exactly match the user\'s display name or login in Odoo.\n'
            '4. Team and Shift names must exactly match existing records in Odoo.',
            note_fmt)
        ws.set_row(13, 50)
        ws.set_row(14, 20)

    def _build_data_sheet(self, workbook):
        ws = workbook.add_worksheet('Import Data')

        req_hdr = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#C55A11', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True,
        })
        cond_hdr = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#9C6500', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True,
        })
        opt_hdr = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#375623', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True,
        })
        sample_fmt = workbook.add_format({
            'align': 'left', 'valign': 'vcenter', 'border': 1,
            'bg_color': '#F2F2F2', 'italic': True, 'font_color': '#7F7F7F',
        })

        columns = [
            ('Team *',                              20, req_hdr),
            ('Distribution Type *\nsingle / round_robin', 24, req_hdr),
            ('Shift Name *',                        18, req_hdr),
            ('Shift Date *\nYYYY-MM-DD',            18, req_hdr),
            ('Single Agent\n(for single mode)',     24, cond_hdr),
            ('RR Agent 1\n(for round_robin)',        22, cond_hdr),
            ('RR Agent 2',                          18, opt_hdr),
            ('RR Agent 3',                          18, opt_hdr),
            ('RR Agent 4',                          18, opt_hdr),
            ('RR Agent 5',                          18, opt_hdr),
            ('Allocation Type\napartment / lands',  20, opt_hdr),
        ]

        ws.set_row(0, 40)
        for col, (header, width, fmt) in enumerate(columns):
            ws.write(0, col, header, fmt)
            ws.set_column(col, col, width)

        # Sample row 1 — single agent
        sample1 = ['Sales Team A', 'single', 'Morning Shift', '2025-06-01',
                   'John Smith', '', '', '', '', '', 'lands']
        for col, val in enumerate(sample1):
            ws.write(1, col, val, sample_fmt)
        ws.set_row(1, 16)

        # Sample row 2 — round robin
        sample2 = ['Sales Team B', 'round_robin', 'Evening Shift', '2025-06-01',
                   '', 'Alice Brown', 'Bob Wilson', 'Carol Davis', '', '', 'apartment']
        for col, val in enumerate(sample2):
            ws.write(2, col, val, sample_fmt)
        ws.set_row(2, 16)

    def action_import(self):
        if not self.import_file:
            raise UserError(_('Please upload an Excel file before importing.'))

        try:
            import openpyxl
        except ImportError:
            raise UserError(_(
                'The "openpyxl" library is required for import. '
                'Install it with: pip install openpyxl'
            ))

        raw = self.with_context(bin_size=False).import_file
        wb = None
        for data in (base64.b64decode(raw), raw):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
                break
            except Exception:
                continue
        if wb is None:
            raise UserError(_(
                'Could not open the uploaded file. '
                'Please make sure you upload a valid .xlsx file created from the template.'
            ))

        if 'Import Data' not in wb.sheetnames:
            raise UserError(_(
                'The uploaded file must contain a sheet named "Import Data". '
                'Please use the provided template.'
            ))

        ws = wb['Import Data']
        errors = []
        created = 0
        SAMPLE_TEAMS = {'Sales Team A', 'Sales Team B'}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            team_name         = str(row[0] or '').strip()
            dist_type         = str(row[1] or '').strip().lower()
            shift_name        = str(row[2] or '').strip()
            shift_date_raw    = row[3]
            single_agent_name = str(row[4] or '').strip()
            rr_names          = [str(row[i] or '').strip() for i in range(5, 10) if str(row[i] or '').strip()]
            alloc_type        = str(row[10] or '').strip().lower() if len(row) > 10 else ''

            if team_name in SAMPLE_TEAMS:
                continue

            # Required field checks
            row_errors = []
            if not team_name:
                row_errors.append('Team is required')
            if dist_type not in ('single', 'round_robin'):
                row_errors.append('Distribution Type must be "single" or "round_robin" (got: "%s")' % dist_type)
            if not shift_name:
                row_errors.append('Shift Name is required')
            if not shift_date_raw:
                row_errors.append('Shift Date is required')
            if alloc_type and alloc_type not in ('apartment', 'lands'):
                row_errors.append('Allocation Type must be "apartment" or "lands" (got: "%s")' % alloc_type)

            if row_errors:
                errors.append('Row %d: %s' % (row_idx, '; '.join(row_errors)))
                continue

            # Parse date
            try:
                if hasattr(shift_date_raw, 'date'):
                    shift_date = shift_date_raw.date()
                else:
                    shift_date = date.fromisoformat(str(shift_date_raw).strip())
            except Exception:
                errors.append('Row %d: Invalid Shift Date "%s" — use YYYY-MM-DD format.' % (row_idx, shift_date_raw))
                continue

            # Look up team
            team = self.env['crm.team'].search([('name', '=ilike', team_name)], limit=1)
            if not team:
                errors.append('Row %d: CRM Team not found: "%s".' % (row_idx, team_name))
                continue

            # Look up shift
            shift = self.env['lead.shift'].search(
                [('name', '=ilike', shift_name), ('active', '=', True)], limit=1
            )
            if not shift:
                errors.append('Row %d: Shift not found: "%s".' % (row_idx, shift_name))
                continue

            # Build allocation values
            alloc_vals = {
                'team_id': team.id,
                'distribution_type': dist_type,
                'shift_id': shift.id,
                'shift_date': shift_date,
                'allocation_type': alloc_type or 'apartment',
            }

            if dist_type == 'single':
                if not single_agent_name:
                    errors.append('Row %d: Single Agent is required for "single" distribution type.' % row_idx)
                    continue
                user = self.env['res.users'].search(
                    ['|', ('name', '=ilike', single_agent_name),
                     ('login', '=ilike', single_agent_name),
                     ('active', '=', True)],
                    limit=1,
                )
                if not user:
                    errors.append('Row %d: User not found: "%s".' % (row_idx, single_agent_name))
                    continue
                alloc_vals['user_id'] = user.id

            else:  # round_robin
                if not rr_names:
                    errors.append('Row %d: At least one Round Robin Agent is required.' % row_idx)
                    continue
                if len(rr_names) > 5:
                    errors.append('Row %d: Maximum 5 Round Robin Agents allowed (got %d).' % (row_idx, len(rr_names)))
                    continue

                line_vals = []
                agent_error = False
                for seq, agent_name in enumerate(rr_names, start=1):
                    u = self.env['res.users'].search(
                        ['|', ('name', '=ilike', agent_name),
                         ('login', '=ilike', agent_name),
                         ('active', '=', True)],
                        limit=1,
                    )
                    if not u:
                        errors.append('Row %d: User not found: "%s".' % (row_idx, agent_name))
                        agent_error = True
                        break
                    line_vals.append((0, 0, {'user_id': u.id, 'sequence': seq * 10}))

                if agent_error:
                    continue
                alloc_vals['allocate_line_ids'] = line_vals

            # Create record
            try:
                self.env['lead.allocate'].create(alloc_vals)
                created += 1
            except Exception as exc:
                errors.append('Row %d: Could not create allocation — %s' % (row_idx, str(exc)))

        # Result reporting
        if not errors and created == 0:
            raise UserError(_(
                'No records were imported. '
                'Make sure to delete the grey sample rows from the template before uploading.'
            ))

        if errors and created == 0:
            raise UserError(_('Import failed — no records created.\n\n') + '\n'.join(errors))

        if errors:
            raise UserError(
                _('%d allocation(s) created, but %d row(s) had errors:\n\n') % (created, len(errors)) +
                '\n'.join(errors)
            )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Successful'),
                'message': _('%d lead allocation(s) successfully imported.') % created,
                'type': 'success',
                'sticky': False,
            },
        }
